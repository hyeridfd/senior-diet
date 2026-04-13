"""
report_agent.py  ─  ReportAgent 노드 (registry 버전)
개인화 대체 메뉴(personal_menus) 시트 추가
"""

import os
import pandas as pd
import registry
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from state import MealPlanState

TYPE_COLORS = {
    "HK형": "FCE4D6", "DHK형": "FCE4D6", "K형": "FCE4D6", "DK형": "FCE4D6",
    "DH형": "FFF2CC", "D형":   "FFF2CC",
    "H형":  "EBF1DE", "일반형": "FFFFFF",
}

SLOT_CATS = [
    ("밥","밥"),("국","국"),("주찬","주찬"),
    ("부찬1","부찬"),("부찬2","부찬"),("김치","김치"),
]


def report_agent(state: MealPlanState) -> dict:
    print("\n[ReportAgent] 보고서 생성 시작...")

    if not state.get("df_menu_records"):
        print("  [경고] df_menu_records 없음 — 건너뜀")
        return {"report_paths": {}, "messages": ["[ReportAgent] df_menu 없음"]}

    df = pd.DataFrame(
        state["df_menu_records"],
        columns=state["df_menu_columns"]
    )

    patients = registry.get(state["patients_key"]) if state.get("patients_key") else []

    paths: dict = {}

    # 1. 식단표 Excel
    meal_path = _save_meal_plan_excel(df, state.get("recommend_map") or {})
    paths["meal_plan"] = meal_path

    # 2. 개인별 배식량 + 개인화 대체 메뉴 Excel (시트 2개)
    serving_path = _save_serving_excel(
        df, patients,
        state.get("serving_map")    or {},
        state.get("constraint_key"),
        state.get("personal_menus") or {},   # ← 추가
    )
    paths["serving"] = serving_path

    # 3. GPT-4o 조리 지침서
    openai_key = os.getenv("OPENAI_API_KEY", "")
    if openai_key:
        guide_path = _save_cooking_guide(df, patients, openai_key)
        paths["cooking"] = guide_path
    else:
        print("  OPENAI_API_KEY 미설정 → 조리 지침서 건너뜀")

    print(f"[ReportAgent] 완료 — {list(paths.keys())}")
    return {
        "report_paths": paths,
        "messages":     [f"[ReportAgent] 보고서 생성 완료: {list(paths.values())}"],
    }


# ── 스타일 헬퍼 ──────────────────────────────────────────────
def _hdr(ws, addr, val, bg="1F497D", fg="FFFFFF", size=10):
    c = ws[addr]
    c.value = val
    c.font  = Font(name="맑은 고딕", bold=True, color=fg, size=size)
    c.fill  = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _wrt(ws, addr, val, bg=None, bold=False, size=9, color="000000", align="center"):
    c = ws[addr]
    c.value = val
    c.font  = Font(name="맑은 고딕", bold=bold, size=size, color=color)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)


def _border(ws, r1, r2, c1, c2):
    thin = Side(style="thin")
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for c in row:
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ── 1. 식단표 Excel ──────────────────────────────────────────
def _save_meal_plan_excel(df: pd.DataFrame, recommend_map: dict,
                          path="식단표_28일.xlsx") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "28일_식단표"
    ws.sheet_view.showGridLines = False

    cols   = ["일차","끼니","밥","국","주찬","부찬1","부찬2","김치",
              "열량(kcal)","나트륨(mg)","단백질(g)","비용(원)",
              "권장재료포함메뉴","권장재료포함수"]
    widths = [6,6,12,14,14,14,14,10,10,10,9,9,40,8]

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    _hdr(ws, "A1", "28일 최적 식단표", size=12)
    ws.row_dimensions[1].height = 28

    for i, h in enumerate(cols, 1):
        _hdr(ws, f"{get_column_letter(i)}2", h, bg="2E5A9C", size=9)
    ws.row_dimensions[2].height = 20

    if "권장재료포함수" not in df.columns:
        df = df.copy()
        df["권장재료포함메뉴"] = "-"
        df["권장재료포함수"]   = 0

    for r_idx, row in enumerate(df.itertuples(), 3):
        d = row._asdict()
        rec_count = d.get("권장재료포함수", 0) or 0
        try:
            rec_count = int(rec_count)
        except (TypeError, ValueError):
            rec_count = 0
        bg = "C6EFCE" if rec_count >= 3 else ("E2EFDA" if rec_count >= 1 else "FFFFFF")

        for i, col in enumerate(cols, 1):
            val = d.get(col, "")
            if col == "권장재료포함수" and rec_count >= 1:
                _wrt(ws, f"{get_column_letter(i)}{r_idx}", val,
                     bg="C6EFCE", bold=True, color="375623")
            elif col == "권장재료포함메뉴" and str(val) != "-":
                _wrt(ws, f"{get_column_letter(i)}{r_idx}", val,
                     bg=bg, color="375623", align="left")
            else:
                _wrt(ws, f"{get_column_letter(i)}{r_idx}", val, bg=bg)
        ws.row_dimensions[r_idx].height = 16

    _border(ws, 2, len(df)+2, 1, len(cols))
    wb.save(path)
    print(f"  식단표_28일.xlsx 저장 완료")
    return path


# ── 2. 개인별 배식량 Excel (개인화 대체 시트 포함) ──────────
def _save_serving_excel(df: pd.DataFrame, patients: list,
                        serving_map: dict, constraint_key: str,
                        personal_menus: dict,
                        path="개인별_배식량.xlsx") -> str:
    constraint = registry.get(constraint_key) if constraint_key and registry.has(constraint_key) else None

    wb  = openpyxl.Workbook()

    # ── 시트 1: 개인별 배식량 ────────────────────────────────
    ws1 = wb.active
    ws1.title = "개인별_배식량"
    ws1.sheet_view.showGridLines = False

    col_hdrs = ["이름","질환유형","일차","끼니","ratio",
                "밥(g)","국(ml)","주찬(g)","부찬1(g)","부찬2(g)","김치(g)",
                "예상열량","예상단백질","예상나트륨","예상탄수화물",
                "열량OK","단백질OK","나트륨OK"]
    widths   = [10,8,6,6,6,7,7,7,7,7,7,8,8,8,9,6,6,6]

    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.merge_cells(f"A1:{get_column_letter(len(col_hdrs))}1")
    _hdr(ws1, "A1", "개인별 배식량 및 예상 영양소", bg="1F497D", size=12)
    ws1.row_dimensions[1].height = 28
    for i, h in enumerate(col_hdrs, 1):
        _hdr(ws1, f"{get_column_letter(i)}2", h, bg="2E5A9C", size=9)

    rows_detail = []
    for _, menu_row in df.iterrows():
        day  = menu_row["일차"]
        meal = menu_row["끼니"]
        for p in patients:
            key_str = f"{p.name}||{day}||{meal}"
            srv     = serving_map.get(key_str, {})
            ratio   = srv.get("ratio", 1.0)

            c     = getattr(p, "constraint", None) or constraint
            e_min = getattr(c, "energy_min", 0)    or 0
            e_max = getattr(c, "energy_max", 9999) or 9999
            p_max = getattr(c, "protein_max", 9999) or 9999
            s_max = getattr(c, "sodium_max", 9999)  or 9999

            ok_e = e_min <= (srv.get("예상열량", 0) or 0) <= e_max
            ok_p = (srv.get("예상단백질", 0) or 0) <= p_max
            ok_s = (srv.get("예상나트륨", 0) or 0) <= s_max

            rows_detail.append([
                p.name,
                getattr(p, "disease_type_label", "-"),
                day, meal,
                round(ratio, 2),
                srv.get("밥",    srv.get("죽", 0)),
                srv.get("국",    0),
                srv.get("주찬",  0),
                srv.get("부찬1", 0),
                srv.get("부찬2", 0),
                srv.get("김치",  0),
                srv.get("예상열량",     0),
                srv.get("예상단백질",   0),
                srv.get("예상나트륨",   0),
                srv.get("예상탄수화물", 0),
                "✅" if ok_e else "⚠️",
                "✅" if ok_p else "⚠️",
                "✅" if ok_s else "⚠️",
            ])

    for r_idx, row_vals in enumerate(rows_detail, 3):
        disease = row_vals[1]
        bg = TYPE_COLORS.get(disease, "FFFFFF")
        for i, val in enumerate(row_vals, 1):
            color = "974706" if str(val) == "⚠️" else "000000"
            _wrt(ws1, f"{get_column_letter(i)}{r_idx}", val,
                 bg=bg, size=9, color=color)
        ws1.row_dimensions[r_idx].height = 14

    _border(ws1, 2, len(rows_detail)+2, 1, len(col_hdrs))

    # ── 시트 2: 개인화 대체 메뉴 (조리팀용) ─────────────────
    if personal_menus:
        ws2 = wb.create_sheet("개인화_부찬대체")
        ws2.sheet_view.showGridLines = False

        p_hdrs  = ["이름", "일차", "끼니", "기존 부찬1", "대체 부찬1", "비고"]
        p_widths = [12, 8, 8, 18, 18, 20]
        for i, w in enumerate(p_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        ws2.merge_cells(f"A1:{get_column_letter(len(p_hdrs))}1")
        _hdr(ws2, "A1", "개인화 부찬 대체 지침 (조리팀용)", bg="375623", size=12)
        ws2.row_dimensions[1].height = 28
        for i, h in enumerate(p_hdrs, 1):
            _hdr(ws2, f"{get_column_letter(i)}2", h, bg="4E7C2F", size=9)
        ws2.row_dimensions[2].height = 20

        # df_menu → 원래 부찬1 조회용 인덱스
        df_idx = {
            (row["일차"], row["끼니"]): row.get("부찬1", "")
            for _, row in df.iterrows()
        }

        # personal_menus 파싱 및 정렬
        p_rows = []
        for key, changes in personal_menus.items():
            name, day, meal = key.split("||")
            orig = df_idx.get((day, meal), "-")
            alt  = changes.get("부찬1", "-")
            p_rows.append([name, day, meal, orig, alt])

        # 이름 → 일차 → 끼니 순 정렬
        meal_order = {"아침": 0, "점심": 1, "저녁": 2}
        p_rows.sort(key=lambda x: (
            x[0],
            int(x[1].replace("일", "")),
            meal_order.get(x[2], 9),
        ))

        patient_map = {getattr(p, "name", ""): p for p in patients}

        for r_idx, row_vals in enumerate(p_rows, 3):
            name    = row_vals[0]
            p       = patient_map.get(name)
            disease = getattr(p, "disease_type_label", "일반형") if p else "일반형"
            bg      = TYPE_COLORS.get(disease, "FFFFFF")

            for i, val in enumerate(row_vals, 1):
                bold = (i == 5)       # 대체 부찬1 컬럼 굵게
                color = "375623" if i == 5 else "000000"
                _wrt(ws2, f"{get_column_letter(i)}{r_idx}", val,
                     bg=bg, bold=bold, size=9, color=color)

            # 비고: 질환 유형 표시
            _wrt(ws2, f"F{r_idx}", f"{disease} 기피메뉴 대체",
                 bg=bg, size=9, color="555555", align="left")
            ws2.row_dimensions[r_idx].height = 16

        _border(ws2, 2, len(p_rows)+2, 1, len(p_hdrs))

        print(f"  개인화_부찬대체 시트 저장 완료 ({len(p_rows)}건)")
    else:
        print("  개인화 대체 없음 — 시트 생략")

    wb.save(path)
    print(f"  개인별_배식량.xlsx 저장 완료 ({len(rows_detail)}행)")
    return path


# ── 3. GPT-4o 조리 지침서 ────────────────────────────────────
def _save_cooking_guide(df: pd.DataFrame, patients: list,
                        api_key: str, path="조리_지침서.txt") -> str:
    from openai import OpenAI

    row = df[(df["일차"] == "1일") & (df["끼니"] == "점심")]
    if row.empty:
        return ""
    row = row.iloc[0]
    menu_summary = " / ".join([
        row["밥"], row["국"], row["주찬"],
        row["부찬1"], row["부찬2"], row["김치"]
    ])
    disease_labels = list({
        getattr(p, "disease_type_label", "일반형") for p in patients
    })

    prompt = f"""
노인요양시설 조리 지침서를 작성해 주세요.
[오늘 메뉴] {menu_summary}
[입소자 질환 유형] {', '.join(disease_labels)}

각 메뉴별로:
1. 조리 시 주의사항 (나트륨, 당, 식감 조절)
2. 질환별 배식 조정 포인트
3. 위생 및 온도 관리

간결하고 실용적으로 작성해 주세요.
"""
    client   = OpenAI(api_key=api_key)
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=600, temperature=0.3,
    )
    guide_text = response.choices[0].message.content.strip()

    with open(path, "w", encoding="utf-8") as f:
        f.write(guide_text)
    print(f"  조리_지침서.txt 저장 완료")
    return path